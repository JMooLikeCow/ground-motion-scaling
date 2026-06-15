"""
Generate the output Excel workbook with native charts (no PNG embedding).
Each plot sheet contains the raw numerical data and a live ScatterChart
that references it directly.
"""

import io
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import ScatterChart, Reference, Series

from core.scaling import ScalingResult
from core.compliance import SuiteCompliance


_GREEN    = "FF92D050"
_RED      = "FFFF0000"
_DARK_RED = "FFC00000"
_AMBER    = "FFFFC000"
_BLUE_HDR = "FF4C72B0"
_WHITE    = "FFFFFFFF"
_BLACK    = "FF000000"


def _hdr_style(ws, row, col, value, bg=_BLUE_HDR):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True, color=_WHITE, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return cell


def _thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _set_line(series, hex_rgb, width_pt=1.0, dash=None):
    """Apply colour and weight to a chart series line."""
    series.graphicalProperties.line.solidFill = hex_rgb
    series.graphicalProperties.line.width = int(width_pt * 12700)
    if dash:
        series.graphicalProperties.line.prstDash = dash


def build_excel(
    scaling_results: dict[str, ScalingResult],
    compliance_results: list[SuiteCompliance],
    report_md: str,
    periods: np.ndarray,
    at2_records: dict,
    sa_target_h: np.ndarray,
    t_min: float,
    t_max: float,
) -> bytes:
    """
    Build the output workbook and return as bytes for Streamlit download.

    Sheets:
      SUMMARY          — scale factors and compliance flags
      REPORT           — design note
      RECORDS_LOG      — per-file AT2 metadata
      SPECTRAL_RATIOS  — numerical Sa(scaled)/Sa(target) table
      PLOT_SPECTRA_FULL  — data + native chart (full period range)
      PLOT_SPECTRA_ZOOM  — data + native chart (period range of interest)
      PLOT_DEV_FULL      — data + native chart (deviation ratio full)
      PLOT_DEV_ZOOM      — data + native chart (deviation ratio zoomed)
    """
    alpha_h = compliance_results[0].alpha_h

    wb = Workbook()
    wb.remove(wb.active)

    _write_summary(wb, scaling_results, compliance_results, periods)
    _write_report(wb, report_md)
    _write_records_log(wb, at2_records, scaling_results)
    _write_spectral_ratios(wb, scaling_results, sa_target_h, periods, t_min, t_max, alpha_h)
    _write_spectra_plot(
        wb, "PLOT_SPECTRA_FULL",
        "Response Spectra — Scaled Suite vs Target (full range)",
        periods, scaling_results, sa_target_h, alpha_h, t_min, t_max,
        zoomed=False,
    )
    _write_spectra_plot(
        wb, "PLOT_SPECTRA_ZOOM",
        "Response Spectra — Scaled Suite vs Target (period range of interest)",
        periods, scaling_results, sa_target_h, alpha_h, t_min, t_max,
        zoomed=True,
    )
    _write_deviation_plot(
        wb, "PLOT_DEV_FULL",
        "Deviation Ratio — Mean Sa / Target Sa (full range)",
        periods, scaling_results, sa_target_h, alpha_h, t_min, t_max,
        zoomed=False,
    )
    _write_deviation_plot(
        wb, "PLOT_DEV_ZOOM",
        "Deviation Ratio — Mean Sa / Target Sa (period range of interest)",
        periods, scaling_results, sa_target_h, alpha_h, t_min, t_max,
        zoomed=True,
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── SUMMARY ───────────────────────────────────────────────────────────────────

def _write_summary(wb, scaling_results, compliance_results, periods):
    ws = wb.create_sheet("SUMMARY")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "GROUND MOTION SCALING — SUMMARY"
    c.font = Font(bold=True, size=13, color=_WHITE)
    c.fill = PatternFill("solid", fgColor=_BLUE_HDR)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    row = 3
    headers = ["Record ID", "k1 (log-space)", "k2 (suite corr.)", "SF (H) final",
               "SF (V) final", "Scaled PGA H1 (g)"]
    for ci, h in enumerate(headers, 1):
        _hdr_style(ws, row, ci, h)

    row += 1
    for rid, r in scaling_results.items():
        ws.cell(row=row, column=1, value=rid)
        ws.cell(row=row, column=2, value=round(r.sf_h_k1, 4))
        ws.cell(row=row, column=3, value=round(r.sf_h_k2, 4))
        ws.cell(row=row, column=4, value=round(r.sf_h, 4))
        ws.cell(row=row, column=5, value=round(r.sf_v, 4) if r.sf_v else "N/A")
        pga_h = r.sf_h * float(np.max(np.abs(r.sa_h1_unscaled))) if r.sa_h1_unscaled is not None else "N/A"
        ws.cell(row=row, column=6, value=round(pga_h, 4) if isinstance(pga_h, float) else pga_h)
        for ci in range(1, 7):
            ws.cell(row=row, column=ci).border = _thin_border()
        row += 1

    row += 1
    for comp in compliance_results:
        _hdr_style(ws, row, 1, f"Compliance — {comp.code}", bg="FF2E5E9B")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
        for label, pass_val, alpha in [
            ("Horizontal suite", comp.suite_pass_h, comp.alpha_h),
            ("Vertical suite",   comp.suite_pass_v, comp.alpha_v),
        ]:
            if pass_val is None:
                continue
            ws.cell(row=row, column=1, value=f"{label} (α = {alpha:.2f})")
            cell = ws.cell(row=row, column=2, value="PASS" if pass_val else "FAIL")
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor=_GREEN if pass_val else _RED)
            row += 1
        if comp.min_records_warning:
            from core.compliance import MIN_RECORDS
            min_r = MIN_RECORDS[comp.code]
            ws.cell(row=row, column=1,
                    value=f"⚠ Record count below {comp.code} minimum ({min_r})")
            ws.cell(row=row, column=1).font = Font(color="FFFF0000")
            row += 1
        row += 1

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 20


# ── REPORT ────────────────────────────────────────────────────────────────────

def _write_report(wb, report_md: str):
    ws = wb.create_sheet("REPORT")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 120
    for i, line in enumerate(report_md.split("\n"), 1):
        cell = ws.cell(row=i, column=1, value=line)
        if line.startswith("## "):
            cell.font = Font(bold=True, size=12, color=_WHITE)
            cell.fill = PatternFill("solid", fgColor=_BLUE_HDR)
        elif line.startswith("### "):
            cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(wrap_text=True)


# ── RECORDS LOG ───────────────────────────────────────────────────────────────

def _write_records_log(wb, at2_records, scaling_results):
    ws = wb.create_sheet("RECORDS_LOG")
    ws.sheet_view.showGridLines = False
    headers = ["Record ID", "Component", "Filename", "NPTS", "DT (s)",
               "Duration (s)", "Unscaled PGA (g)", "Scale Factor Applied"]
    for ci, h in enumerate(headers, 1):
        _hdr_style(ws, 1, ci, h)
    row = 2
    for rid, group in at2_records.items():
        sf_h = scaling_results[rid].sf_h if rid in scaling_results else None
        sf_v = scaling_results[rid].sf_v if rid in scaling_results else None
        for comp, rec in group.items():
            sf = sf_v if comp == "V" else sf_h
            ws.cell(row=row, column=1, value=rid)
            ws.cell(row=row, column=2, value=comp)
            ws.cell(row=row, column=3, value=rec.filename)
            ws.cell(row=row, column=4, value=rec.npts)
            ws.cell(row=row, column=5, value=rec.dt)
            ws.cell(row=row, column=6, value=round(rec.duration, 3))
            ws.cell(row=row, column=7, value=round(rec.pga, 5))
            ws.cell(row=row, column=8, value=round(sf, 4) if sf else "N/A")
            for ci in range(1, 9):
                ws.cell(row=row, column=ci).border = _thin_border()
            row += 1
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 22


# ── SPECTRAL RATIOS ───────────────────────────────────────────────────────────

def _write_spectral_ratios(wb, scaling_results, sa_target_h, periods, t_min, t_max, alpha_h):
    ws = wb.create_sheet("SPECTRAL_RATIOS")
    ws.sheet_view.showGridLines = False

    mask = (periods >= t_min) & (periods <= t_max)
    p = periods[mask]
    tgt = sa_target_h[mask]
    record_ids = list(scaling_results.keys())

    all_scaled = np.vstack([scaling_results[rid].sa_combined_scaled[mask] for rid in record_ids])
    mean_scaled = np.mean(all_scaled, axis=0)
    mean_ratios = np.where(tgt > 0, mean_scaled / tgt, np.nan)

    # Headers: Period | Suite Mean | Record1 | Record2 | ...
    headers = ["Period (s)", "Suite Mean"] + record_ids
    for ci, h in enumerate(headers, 1):
        c = _hdr_style(ws, 1, ci, h)

    for ri, (t_val, mean_r) in enumerate(zip(p, mean_ratios), 2):
        ws.cell(row=ri, column=1, value=round(float(t_val), 3))
        # Suite Mean with colour
        mean_cell = ws.cell(row=ri, column=2, value=round(float(mean_r), 3))
        if mean_r < alpha_h:
            mean_cell.fill = PatternFill("solid", fgColor="FFCCCC")
        elif mean_r < 1.0:
            mean_cell.fill = PatternFill("solid", fgColor="FFF3CC")
        # Individual records
        for ci, rid in enumerate(record_ids, 3):
            sa_rec = scaling_results[rid].sa_combined_scaled[mask][ri - 2]
            ratio = float(sa_rec) / float(tgt[ri - 2]) if tgt[ri - 2] > 0 else float("nan")
            ws.cell(row=ri, column=ci, value=round(ratio, 3))

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16


# ── SPECTRA PLOT SHEET ────────────────────────────────────────────────────────

def _write_spectra_plot(wb, sheet_name, title, periods, scaling_results,
                        sa_target_h, alpha_h, t_min, t_max, zoomed):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    if zoomed:
        pad = (t_max - t_min) * 0.05
        mask = (periods >= t_min - pad) & (periods <= t_max + pad)
    else:
        mask = np.ones(len(periods), dtype=bool)

    p = periods[mask]
    tgt = sa_target_h[mask]
    record_ids = list(scaling_results.keys())

    all_sa = np.vstack([scaling_results[rid].sa_combined_scaled[mask] for rid in record_ids])
    mean_sa = np.mean(all_sa, axis=0)

    # Columns: Period | Rec1 | Rec2 | ... | Suite Mean | Target | α×Target
    suite_col = len(record_ids) + 2
    target_col = suite_col + 1
    alpha_target_col = suite_col + 2

    headers = (["Period (s)"] + record_ids +
               ["Suite Mean", "Target", f"α×Target (α={alpha_h:.2f})"])
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h).font = Font(bold=True)

    n_rows = len(p)
    for ri, (t_val, mean_v, tgt_v) in enumerate(zip(p, mean_sa, tgt), 2):
        ws.cell(row=ri, column=1, value=round(float(t_val), 4))
        for ci, rid in enumerate(record_ids, 2):
            ws.cell(row=ri, column=ci,
                    value=round(float(scaling_results[rid].sa_combined_scaled[mask][ri - 2]), 4))
        ws.cell(row=ri, column=suite_col,        value=round(float(mean_v), 4))
        ws.cell(row=ri, column=target_col,       value=round(float(tgt_v), 4))
        ws.cell(row=ri, column=alpha_target_col, value=round(float(alpha_h * tgt_v), 4))

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    # ── Native ScatterChart ───────────────────────────────────────────────────
    chart = ScatterChart()
    chart.scatterStyle = "smooth"
    chart.title = title
    chart.y_axis.title = "Sa (g)"
    chart.x_axis.title = "Period (s)"
    chart.width = 25
    chart.height = 16

    xref = Reference(ws, min_col=1, min_row=2, max_row=n_rows + 1)

    # Individual records — auto-colored by Excel, thin
    for ci, rid in enumerate(record_ids, 2):
        yref = Reference(ws, min_col=ci, min_row=1, max_row=n_rows + 1)
        s = Series(yref, xref, title_from_data=True)
        chart.series.append(s)

    # Suite Mean — black, 2 pt
    yref = Reference(ws, min_col=suite_col, min_row=1, max_row=n_rows + 1)
    s = Series(yref, xref, title_from_data=True)
    _set_line(s, "000000", width_pt=2.0)
    chart.series.append(s)

    # Target — red solid, 1.5 pt
    yref = Reference(ws, min_col=target_col, min_row=1, max_row=n_rows + 1)
    s = Series(yref, xref, title_from_data=True)
    _set_line(s, "FF0000", width_pt=1.5)
    chart.series.append(s)

    # α×Target — dark red dashed, 1 pt
    yref = Reference(ws, min_col=alpha_target_col, min_row=1, max_row=n_rows + 1)
    s = Series(yref, xref, title_from_data=True)
    _set_line(s, _DARK_RED.lstrip("FF"), width_pt=1.0, dash="dash")
    chart.series.append(s)

    ws.add_chart(chart, f"A{n_rows + 3}")


# ── DEVIATION RATIO PLOT SHEET ────────────────────────────────────────────────

def _write_deviation_plot(wb, sheet_name, title, periods, scaling_results,
                          sa_target_h, alpha_h, t_min, t_max, zoomed):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    if zoomed:
        pad = (t_max - t_min) * 0.05
        mask = (periods >= t_min - pad) & (periods <= t_max + pad)
    else:
        mask = np.ones(len(periods), dtype=bool)

    p = periods[mask]
    tgt = sa_target_h[mask]
    record_ids = list(scaling_results.keys())

    all_sa = np.vstack([scaling_results[rid].sa_combined_scaled[mask] for rid in record_ids])
    mean_sa = np.mean(all_sa, axis=0)
    with np.errstate(divide="ignore", invalid="ignore"):
        ratio = np.where(tgt > 0, mean_sa / tgt, np.nan)

    # Columns: Period | Mean/Target | 1.0 (threshold) | α (compliance)
    headers = ["Period (s)", "Mean Sa / Target Sa", "1.0 (target)", f"α = {alpha_h:.2f} (compliance)"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h).font = Font(bold=True)

    n_rows = len(p)
    for ri, (t_val, r_val) in enumerate(zip(p, ratio), 2):
        ws.cell(row=ri, column=1, value=round(float(t_val), 4))
        ws.cell(row=ri, column=2, value=round(float(r_val), 4) if np.isfinite(r_val) else None)
        ws.cell(row=ri, column=3, value=1.0)
        ws.cell(row=ri, column=4, value=round(alpha_h, 2))

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 22

    # ── Native ScatterChart ───────────────────────────────────────────────────
    chart = ScatterChart()
    chart.scatterStyle = "smooth"
    chart.title = title
    chart.y_axis.title = "Mean Sa / Target Sa"
    chart.x_axis.title = "Period (s)"
    chart.width = 25
    chart.height = 16

    xref = Reference(ws, min_col=1, min_row=2, max_row=n_rows + 1)

    # Mean ratio — blue, 2 pt
    yref = Reference(ws, min_col=2, min_row=1, max_row=n_rows + 1)
    s = Series(yref, xref, title_from_data=True)
    _set_line(s, "1F77B4", width_pt=2.0)
    chart.series.append(s)

    # 1.0 reference — red solid, 1 pt
    yref = Reference(ws, min_col=3, min_row=1, max_row=n_rows + 1)
    s = Series(yref, xref, title_from_data=True)
    _set_line(s, "FF0000", width_pt=1.0)
    chart.series.append(s)

    # α reference — orange dashed, 1 pt
    yref = Reference(ws, min_col=4, min_row=1, max_row=n_rows + 1)
    s = Series(yref, xref, title_from_data=True)
    _set_line(s, "FF8C00", width_pt=1.0, dash="dash")
    chart.series.append(s)

    ws.add_chart(chart, f"A{n_rows + 3}")
